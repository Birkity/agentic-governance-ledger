from __future__ import annotations

import csv
import hashlib
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

from src.document_processing.models import DocumentPackageAnalysis, DocumentPartResult
from src.event_store import EventStore, InMemoryEventStore
from src.models.events import (
    DocumentAdded,
    DocumentFormatValidated,
    DocumentType,
    ExtractionCompleted,
    ExtractionStarted,
    FinancialFacts,
    PackageCreated,
    PackageReadyForAnalysis,
    QualityAssessmentCompleted,
)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _document_id(document: DocumentPartResult) -> str:
    return f"{document.company_id}:{document.document_type.value}"


def _hash_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _count_pages_or_units(document: DocumentPartResult) -> int:
    path = Path(document.path)
    if document.document_format.value == "pdf":
        return len(PdfReader(str(path)).pages)
    if document.document_format.value == "xlsx":
        return len(load_workbook(path, read_only=True).sheetnames)
    if document.document_format.value == "csv":
        with path.open("r", encoding="utf-8", newline="") as handle:
            return max(sum(1 for _ in csv.reader(handle)) - 1, 1)
    return 1


def _has_facts(facts: FinancialFacts) -> bool:
    return any(value is not None for value in facts.model_dump(exclude_none=True).values())


def _critical_missing_fields(document: DocumentPartResult) -> list[str]:
    fields: list[str] = []
    for note in document.extraction_notes:
        lowered = note.lower()
        if lowered.startswith("missing "):
            field_name = lowered.replace("missing ", "", 1).split(" in ", 1)[0].strip()
            fields.append(field_name)
    return fields


def _quality_result(
    document: DocumentPartResult,
    package_analysis: DocumentPackageAnalysis,
) -> dict[str, Any]:
    critical_missing_fields = _critical_missing_fields(document)
    anomalies = [note for note in document.extraction_notes if note not in {f"Missing {field} in income statement PDF" for field in critical_missing_fields}]
    anomalies.extend(
        note
        for note in package_analysis.consistency_notes
        if document.document_type.value.split("_")[0] in note.lower()
    )

    if package_analysis.package_summary:
        auditor_notes = package_analysis.package_summary
    elif document.summary:
        auditor_notes = document.summary
    else:
        auditor_notes = (
            f"Parsed with {document.parser_used}. "
            "Quality assessment is deterministic and based on extraction notes plus package consistency checks."
        )

    if critical_missing_fields and document.document_type in {DocumentType.INCOME_STATEMENT, DocumentType.BALANCE_SHEET}:
        confidence = 0.55
    elif anomalies:
        confidence = 0.7
    else:
        confidence = 0.95

    return {
        "overall_confidence": confidence,
        "is_coherent": not anomalies,
        "anomalies": anomalies,
        "critical_missing_fields": critical_missing_fields,
        "reextraction_recommended": bool(critical_missing_fields or anomalies),
        "auditor_notes": auditor_notes,
    }


def build_document_package_events(
    package_analysis: DocumentPackageAnalysis,
    application_id: str,
    package_id: str | None = None,
    pipeline_version: str = "phase1-document-processor",
) -> list[dict]:
    package_id = package_id or f"docpkg-{application_id}"
    created_at = _now()

    events: list[dict] = [
        PackageCreated(
            package_id=package_id,
            application_id=application_id,
            required_documents=[document.document_type for document in package_analysis.documents],
            created_at=created_at,
        ).to_store_dict()
    ]

    for document in package_analysis.documents:
        document_id = _document_id(document)
        page_count = _count_pages_or_units(document)
        extraction_started_at = _now()
        quality = _quality_result(document, package_analysis)
        has_facts = _has_facts(document.financial_facts)

        document_added = DocumentAdded(
            package_id=package_id,
            document_id=document_id,
            document_type=document.document_type,
            document_format=document.document_format,
            file_hash=_hash_file(document.path),
            added_at=_now(),
        ).to_store_dict()
        document_added["metadata"] = {
            "document_path": document.path,
            "parser_used": document.parser_used,
        }
        events.append(document_added)

        validated = DocumentFormatValidated(
            package_id=package_id,
            document_id=document_id,
            document_type=document.document_type,
            page_count=page_count,
            detected_format=document.document_format.value,
            validated_at=_now(),
        ).to_store_dict()
        validated["metadata"] = {
            "parser_used": document.parser_used,
        }
        events.append(validated)

        started = ExtractionStarted(
            package_id=package_id,
            document_id=document_id,
            document_type=document.document_type,
            pipeline_version=pipeline_version,
            extraction_model=document.parser_used,
            started_at=extraction_started_at,
        ).to_store_dict()
        events.append(started)

        completed = ExtractionCompleted(
            package_id=package_id,
            document_id=document_id,
            document_type=document.document_type,
            facts=document.financial_facts if has_facts else None,
            raw_text_length=len(document.extracted_text),
            tables_extracted=1 if document.document_format.value in {"xlsx", "csv"} else 0,
            processing_ms=0,
            completed_at=_now(),
        ).to_store_dict()
        completed["metadata"] = {
            "document_path": document.path,
            "parser_used": document.parser_used,
            "structured_data": document.structured_data,
            "document_summary": document.summary,
            "extraction_notes": document.extraction_notes,
        }
        events.append(completed)

        quality_event = QualityAssessmentCompleted(
            package_id=package_id,
            document_id=document_id,
            overall_confidence=quality["overall_confidence"],
            is_coherent=quality["is_coherent"],
            anomalies=quality["anomalies"],
            critical_missing_fields=quality["critical_missing_fields"],
            reextraction_recommended=quality["reextraction_recommended"],
            auditor_notes=quality["auditor_notes"],
            assessed_at=_now(),
        ).to_store_dict()
        quality_event["metadata"] = {
            "document_type": document.document_type.value,
            "document_summary": document.summary,
        }
        events.append(quality_event)

    total_quality_flags = sum(
        len(_quality_result(document, package_analysis)["critical_missing_fields"])
        + len(_quality_result(document, package_analysis)["anomalies"])
        for document in package_analysis.documents
    )
    package_ready = PackageReadyForAnalysis(
        package_id=package_id,
        application_id=application_id,
        documents_processed=len(package_analysis.documents),
        has_quality_flags=total_quality_flags > 0,
        quality_flag_count=total_quality_flags,
        ready_at=_now(),
    ).to_store_dict()
    package_ready["metadata"] = {
        "field_sources": package_analysis.field_sources,
        "consistency_notes": package_analysis.consistency_notes,
        "package_summary": package_analysis.package_summary,
    }
    events.append(package_ready)
    return events


async def persist_document_package(
    store: EventStore | InMemoryEventStore,
    package_analysis: DocumentPackageAnalysis,
    application_id: str,
    package_id: str | None = None,
    pipeline_version: str = "phase1-document-processor",
) -> list[int]:
    stream_id = f"docpkg-{application_id}"
    expected_version = await store.stream_version(stream_id)
    events = build_document_package_events(
        package_analysis=package_analysis,
        application_id=application_id,
        package_id=package_id or stream_id,
        pipeline_version=pipeline_version,
    )
    return await store.append(stream_id, events, expected_version=expected_version)
