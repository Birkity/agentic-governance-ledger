from __future__ import annotations

import csv
import re
from decimal import Decimal
from pathlib import Path
from typing import Any

import pdfplumber
from openpyxl import load_workbook
from pypdf import PdfReader

from src.document_processing.models import DocumentPackageAnalysis, DocumentPartResult
from src.document_processing.summarizer import BaseSummarizer
from src.models.events import DocumentFormat, DocumentType, FinancialFacts


_MONEY_PATTERNS: dict[str, tuple[str, ...]] = {
    "total_revenue": (r"Revenue\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "gross_profit": (r"Gross Profit\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "operating_expenses": (r"Operating Expenses\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "operating_income": (r"Operating Income \(EBIT\)\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "ebitda": (r"EBITDA\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "depreciation_amortization": (r"Depreciation\s*&\s*Amortization\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "interest_expense": (r"Interest Expense\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "income_before_tax": (r"Income Before Tax\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "tax_expense": (r"Income Tax Expense\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "net_income": (r"Net Income\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "cash_and_equivalents": (r"Cash (?:and Cash )?Equivalents\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "accounts_receivable": (r"Accounts Receivable(?:,\s*net)?\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "inventory": (r"Inventory\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "current_assets": (r"Total Current Assets\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "total_assets": (r"TOTAL ASSETS\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "current_liabilities": (r"Total Current Liabilities\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "long_term_debt": (r"Long-Term Debt\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "total_liabilities": (r"TOTAL LIABILITIES\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "total_equity": (r"TOTAL EQUITY\s+(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
}

_PROPOSAL_PATTERNS: dict[str, tuple[str, ...]] = {
    "legal_entity_name": (r"Legal Entity Name:\s*(.+)",),
    "business_type": (r"Business Type:\s*(.+)",),
    "industry": (r"Industry:\s*(.+)",),
    "jurisdiction": (r"Jurisdiction:\s*([A-Z]{2})",),
    "requested_amount_usd": (r"Requested Amount:\s*(\(?\$?[0-9,]+(?:\.[0-9]+)?\)?)",),
    "purpose": (r"Purpose:\s*(.+)",),
    "use_of_proceeds": (r"Use of Proceeds:\s*(.+)",),
}


def _parse_decimal(raw_value: Any) -> Decimal | None:
    if raw_value is None or raw_value == "":
        return None
    if isinstance(raw_value, Decimal):
        return raw_value
    if isinstance(raw_value, (int, float)):
        return Decimal(str(raw_value))

    text = str(raw_value).strip()
    if not text:
        return None

    negative = text.startswith("(") and text.endswith(")")
    cleaned = text.replace("$", "").replace(",", "").replace("(", "").replace(")", "").replace("%", "").replace("x", "")
    value = Decimal(cleaned)
    return -value if negative else value


def _parse_bool(raw_value: Any) -> bool | None:
    if raw_value is None or raw_value == "":
        return None
    text = str(raw_value).strip().lower()
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return None


def _extract_first(text: str, patterns: tuple[str, ...]) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE | re.MULTILINE)
        if match:
            return match.group(1).strip()
    return None


def _try_docling_extract(path: Path) -> str | None:
    try:
        from docling.document_converter import DocumentConverter
    except ImportError:
        return None

    try:
        converter = DocumentConverter()
        result = converter.convert(str(path))
        document = getattr(result, "document", None)
        if document is None:
            return None
        if hasattr(document, "export_to_markdown"):
            text = document.export_to_markdown()
        elif hasattr(document, "export_to_text"):
            text = document.export_to_text()
        else:
            text = str(document)
        return text.strip() or None
    except Exception:
        return None


def _extract_pdf_text(path: Path) -> tuple[str, str]:
    docling_text = _try_docling_extract(path)
    if docling_text:
        return docling_text, "docling"

    try:
        with pdfplumber.open(path) as pdf:
            text = "\n".join(page.extract_text() or "" for page in pdf.pages).strip()
        if text:
            return text, "pdfplumber"
    except Exception:
        text = ""

    reader = PdfReader(str(path))
    fallback_text = "\n".join(page.extract_text() or "" for page in reader.pages).strip()
    return fallback_text, "pypdf"


def _derive_ratios(facts: FinancialFacts) -> None:
    if facts.total_revenue and facts.gross_profit and facts.gross_margin is None:
        facts.gross_margin = float(facts.gross_profit / facts.total_revenue)
    if facts.total_revenue and facts.ebitda and facts.ebitda_margin is None:
        facts.ebitda_margin = float(facts.ebitda / facts.total_revenue)
    if facts.total_revenue and facts.net_income and facts.net_margin is None:
        facts.net_margin = float(facts.net_income / facts.total_revenue)
    if facts.total_equity not in {None, Decimal("0")} and facts.total_liabilities is not None and facts.debt_to_equity is None:
        facts.debt_to_equity = float(facts.total_liabilities / facts.total_equity)
    if facts.current_liabilities not in {None, Decimal("0")} and facts.current_assets is not None and facts.current_ratio is None:
        facts.current_ratio = float(facts.current_assets / facts.current_liabilities)
    if facts.ebitda not in {None, Decimal("0")} and facts.long_term_debt is not None and facts.debt_to_ebitda is None:
        facts.debt_to_ebitda = float(facts.long_term_debt / facts.ebitda)
    if facts.interest_expense not in {None, Decimal("0")} and facts.operating_income is not None and facts.interest_coverage is None:
        facts.interest_coverage = float(facts.operating_income / facts.interest_expense)
    if facts.total_assets is not None and facts.total_liabilities is not None and facts.total_equity is not None:
        discrepancy = facts.total_assets - (facts.total_liabilities + facts.total_equity)
        facts.balance_discrepancy_usd = discrepancy
        facts.balance_sheet_balances = abs(discrepancy) <= Decimal("500")


def _material_difference(left: Any, right: Any) -> bool:
    if left is None or right is None:
        return False
    if isinstance(left, bool) and isinstance(right, bool):
        return left != right

    left_decimal = _parse_decimal(left)
    right_decimal = _parse_decimal(right)
    if left_decimal is None or right_decimal is None:
        return left != right

    tolerance = Decimal("0.05") if abs(left_decimal) < 10 and abs(right_decimal) < 10 else Decimal("5000")
    return abs(left_decimal - right_decimal) > tolerance


class DocumentPackageProcessor:
    """Read generated company documents and normalize them into one package analysis."""

    def __init__(self, documents_root: str | Path = "documents", summarizer: BaseSummarizer | None = None):
        self.documents_root = Path(documents_root)
        self.summarizer = summarizer

    def process_company(self, company_id: str, include_summaries: bool = True) -> DocumentPackageAnalysis:
        company_dir = self.documents_root / company_id
        if not company_dir.exists():
            raise FileNotFoundError(f"Document directory not found: {company_dir}")

        documents = [
            self._parse_income_statement(company_id, company_dir / "income_statement_2024.pdf"),
            self._parse_balance_sheet(company_id, company_dir / "balance_sheet_2024.pdf"),
            self._parse_application_proposal(company_id, company_dir / "application_proposal.pdf"),
            self._parse_excel_workbook(company_id, company_dir / "financial_statements.xlsx"),
            self._parse_csv_summary(company_id, company_dir / "financial_summary.csv"),
        ]

        merged_facts, field_sources = self._merge_facts(documents)
        consistency_notes = self._build_consistency_notes(documents, merged_facts)

        analysis = DocumentPackageAnalysis(
            company_id=company_id,
            documents=documents,
            merged_financial_facts=merged_facts,
            field_sources=field_sources,
            consistency_notes=consistency_notes,
        )

        if include_summaries and self.summarizer is not None:
            summarized_documents = []
            for document in analysis.documents:
                summary = self.summarizer.summarize_document(document)
                summarized_documents.append(document.model_copy(update={"summary": summary}))
            analysis = analysis.model_copy(update={"documents": summarized_documents})
            analysis.package_summary = self.summarizer.summarize_package(analysis)

        return analysis

    def _parse_income_statement(self, company_id: str, path: Path) -> DocumentPartResult:
        text, parser_used = _extract_pdf_text(path)
        facts = FinancialFacts()
        notes: list[str] = []

        for field_name in (
            "total_revenue",
            "gross_profit",
            "operating_expenses",
            "operating_income",
            "ebitda",
            "depreciation_amortization",
            "interest_expense",
            "income_before_tax",
            "tax_expense",
            "net_income",
        ):
            raw_value = _extract_first(text, _MONEY_PATTERNS[field_name])
            value = _parse_decimal(raw_value)
            if value is not None and field_name in {
                "operating_expenses",
                "depreciation_amortization",
                "interest_expense",
                "tax_expense",
            }:
                value = abs(value)
            setattr(facts, field_name, value)
            if getattr(facts, field_name) is None:
                notes.append(f"Missing {field_name} in income statement PDF")

        _derive_ratios(facts)
        return DocumentPartResult(
            company_id=company_id,
            document_type=DocumentType.INCOME_STATEMENT,
            document_format=DocumentFormat.PDF,
            path=str(path),
            parser_used=parser_used,
            extracted_text=text,
            financial_facts=facts,
            extraction_notes=notes,
        )

    def _parse_balance_sheet(self, company_id: str, path: Path) -> DocumentPartResult:
        text, parser_used = _extract_pdf_text(path)
        facts = FinancialFacts()
        notes: list[str] = []

        for field_name in (
            "cash_and_equivalents",
            "accounts_receivable",
            "inventory",
            "current_assets",
            "total_assets",
            "current_liabilities",
            "long_term_debt",
            "total_liabilities",
            "total_equity",
        ):
            raw_value = _extract_first(text, _MONEY_PATTERNS[field_name])
            setattr(facts, field_name, _parse_decimal(raw_value))
            if getattr(facts, field_name) is None:
                notes.append(f"Missing {field_name} in balance sheet PDF")

        _derive_ratios(facts)
        if facts.balance_sheet_balances is False:
            notes.append(
                f"Balance sheet discrepancy detected: {facts.balance_discrepancy_usd}"
            )

        return DocumentPartResult(
            company_id=company_id,
            document_type=DocumentType.BALANCE_SHEET,
            document_format=DocumentFormat.PDF,
            path=str(path),
            parser_used=parser_used,
            extracted_text=text,
            financial_facts=facts,
            extraction_notes=notes,
        )

    def _parse_application_proposal(self, company_id: str, path: Path) -> DocumentPartResult:
        text, parser_used = _extract_pdf_text(path)
        structured_data: dict[str, Any] = {}
        notes: list[str] = []

        for field_name, patterns in _PROPOSAL_PATTERNS.items():
            raw_value = _extract_first(text, patterns)
            if raw_value is None:
                notes.append(f"Missing {field_name} in application proposal PDF")
                continue
            if field_name == "requested_amount_usd":
                structured_data[field_name] = _parse_decimal(raw_value)
            else:
                structured_data[field_name] = raw_value

        return DocumentPartResult(
            company_id=company_id,
            document_type=DocumentType.APPLICATION_PROPOSAL,
            document_format=DocumentFormat.PDF,
            path=str(path),
            parser_used=parser_used,
            extracted_text=text,
            structured_data=structured_data,
            extraction_notes=notes,
        )

    def _parse_excel_workbook(self, company_id: str, path: Path) -> DocumentPartResult:
        workbook = load_workbook(path, data_only=True)
        facts = FinancialFacts()
        notes: list[str] = []

        income_mapping = {
            "Revenue": "total_revenue",
            "Gross Profit": "gross_profit",
            "Operating Expenses": "operating_expenses",
            "Depreciation & Amortization": "depreciation_amortization",
            "Operating Income (EBIT)": "operating_income",
            "EBITDA": "ebitda",
            "Interest Expense": "interest_expense",
            "Income Before Tax": "income_before_tax",
            "Income Tax Expense": "tax_expense",
            "Net Income": "net_income",
        }
        balance_mapping = {
            "Cash & Equivalents": "cash_and_equivalents",
            "Accounts Receivable": "accounts_receivable",
            "Inventory": "inventory",
            "Total Current Assets": "current_assets",
            "TOTAL ASSETS": "total_assets",
            "Total Current Liabilities": "current_liabilities",
            "Long-Term Debt": "long_term_debt",
            "TOTAL LIABILITIES": "total_liabilities",
            "TOTAL EQUITY": "total_equity",
        }
        ratio_mapping = {
            "Current Ratio": "current_ratio",
            "Debt/Equity": "debt_to_equity",
            "Debt/EBITDA": "debt_to_ebitda",
            "Interest Coverage": "interest_coverage",
        }

        for sheet_name, mapping in (
            ("Income Statement", income_mapping),
            ("Balance Sheet", balance_mapping),
            ("Key Ratios", ratio_mapping),
        ):
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                label = str(row[0]).strip() if row and row[0] is not None else ""
                normalized_label = label.lstrip()
                if normalized_label not in mapping:
                    continue
                field_name = mapping[normalized_label]
                value = row[-1]
                if field_name in {"current_ratio", "debt_to_equity", "debt_to_ebitda", "interest_coverage"}:
                    setattr(facts, field_name, float(value) if value is not None else None)
                else:
                    setattr(facts, field_name, _parse_decimal(value))

        _derive_ratios(facts)
        if facts.total_revenue is None:
            notes.append("Workbook did not expose FY2024 revenue")

        return DocumentPartResult(
            company_id=company_id,
            document_type=DocumentType.FINANCIAL_WORKBOOK,
            document_format=DocumentFormat.XLSX,
            path=str(path),
            parser_used="openpyxl",
            extracted_text="",
            financial_facts=facts,
            extraction_notes=notes,
        )

    def _parse_csv_summary(self, company_id: str, path: Path) -> DocumentPartResult:
        facts = FinancialFacts()
        notes: list[str] = []

        with path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                field_name = row["field"]
                raw_value = row["value"]

                if field_name == "balance_sheet_check":
                    facts.balance_sheet_balances = _parse_bool(raw_value)
                    continue
                if field_name in {"debt_to_equity", "current_ratio", "debt_to_ebitda", "interest_coverage_ratio", "gross_margin", "ebitda_margin", "net_margin"}:
                    mapped_name = "interest_coverage" if field_name == "interest_coverage_ratio" else field_name
                    setattr(facts, mapped_name, float(raw_value))
                    continue
                if not hasattr(facts, field_name):
                    continue
                current_value = getattr(facts, field_name)
                if isinstance(current_value, float):
                    setattr(facts, field_name, float(raw_value))
                else:
                    setattr(facts, field_name, _parse_decimal(raw_value))

        _derive_ratios(facts)
        if facts.total_revenue is None:
            notes.append("CSV summary missing total_revenue")

        return DocumentPartResult(
            company_id=company_id,
            document_type=DocumentType.FINANCIAL_SUMMARY,
            document_format=DocumentFormat.CSV,
            path=str(path),
            parser_used="csv",
            extracted_text="",
            financial_facts=facts,
            extraction_notes=notes,
        )

    def _merge_facts(self, documents: list[DocumentPartResult]) -> tuple[FinancialFacts, dict[str, str]]:
        merged = FinancialFacts()
        field_sources: dict[str, str] = {}
        preferred_documents = [
            DocumentType.INCOME_STATEMENT,
            DocumentType.BALANCE_SHEET,
            DocumentType.FINANCIAL_WORKBOOK,
            DocumentType.FINANCIAL_SUMMARY,
        ]
        ordered_documents = sorted(
            documents,
            key=lambda document: preferred_documents.index(document.document_type)
            if document.document_type in preferred_documents
            else len(preferred_documents),
        )

        for document in ordered_documents:
            facts = document.financial_facts
            for field_name in FinancialFacts.model_fields:
                current_value = getattr(merged, field_name)
                next_value = getattr(facts, field_name)
                if current_value is None and next_value is not None:
                    setattr(merged, field_name, next_value)
                    field_sources[field_name] = document.document_type.value

        _derive_ratios(merged)
        return merged, field_sources

    def _build_consistency_notes(
        self,
        documents: list[DocumentPartResult],
        merged_facts: FinancialFacts,
    ) -> list[str]:
        notes: list[str] = []
        comparable_documents = [document for document in documents if document.financial_facts != FinancialFacts()]

        for field_name in (
            "total_revenue",
            "ebitda",
            "net_income",
            "total_assets",
            "total_liabilities",
            "total_equity",
        ):
            seen: list[tuple[str, Any]] = []
            for document in comparable_documents:
                value = getattr(document.financial_facts, field_name)
                if value is not None:
                    seen.append((document.document_type.value, value))
            for index, (left_source, left_value) in enumerate(seen):
                for right_source, right_value in seen[index + 1 :]:
                    if _material_difference(left_value, right_value):
                        notes.append(
                            f"{field_name} differs materially between {left_source} and {right_source}"
                        )

        if merged_facts.balance_sheet_balances is False:
            notes.append(
                f"Balance sheet does not fully balance; discrepancy={merged_facts.balance_discrepancy_usd}"
            )

        if merged_facts.ebitda is None:
            notes.append("EBITDA is missing from the merged facts set")

        return notes
